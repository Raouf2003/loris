from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId, errors as bson_errors
from database import db
from io import BytesIO
import openpyxl
import pytz
from datetime import datetime
from fpdf import FPDF

import calendar

TZ = pytz.timezone("Africa/Algiers")

router = APIRouter(prefix="/api/reports", tags=["reports"])

def _today():
    n = datetime.now(TZ)
    return f"{n.year}-{n.month:02d}-{n.day:02d}"

def _hours_diff(start: str | None, end: str | None) -> float:
    if not start or not end:
        return 0.0
    sh, sm = map(int, start.split(":"))
    eh, em = map(int, end.split(":"))
    return round((eh * 60 + em - sh * 60 - sm) / 60, 2)

def _hours_to_hms(h: float) -> str:
    total = int(round(h * 60))
    hh = total // 60
    mm = total % 60
    return f"{hh:02d}:{mm:02d}:00"

async def _enrich(records):
    result = []
    for r in records:
        emp = None
        try:
            emp = await db.employees.find_one({"_id": ObjectId(r["employeeId"])})
        except (bson_errors.InvalidId, TypeError):
            pass
        am_h = _hours_diff(r.get("checkInAM"), r.get("checkOutAM"))
        pm_h = _hours_diff(r.get("checkInPM"), r.get("checkOutPM"))
        result.append({
            "employeeId": {
                "_id": r["employeeId"],
                "fullName": emp["fullName"] if emp else "N/A",
                "employeeNumber": emp["employeeNumber"] if emp else "N/A"
            },
            "date": r["date"],
            "checkInAM": r.get("checkInAM"),
            "checkOutAM": r.get("checkOutAM"),
            "checkInPM": r.get("checkInPM"),
            "checkOutPM": r.get("checkOutPM"),
            "amHours": am_h,
            "pmHours": pm_h,
            "totalHours": round(am_h + pm_h, 2),
            "status": r.get("status", "absent"),
        })
    return result

@router.get("/daily")
async def daily_report(date: str = None):
    date = date or _today()
    records = await db.attendances.find({"date": date}).to_list(None)
    return {"date": date, "records": await _enrich(records)}

@router.get("/monthly")
async def monthly_report(month: int = None, year: int = None):
    n = datetime.now(TZ)
    month = month or n.month
    year = year or n.year
    last_day = calendar.monthrange(year, month)[1]
    start = f"{year:04d}-{month:02d}-01"
    end = f"{year:04d}-{month:02d}-{last_day:02d}"

    records = await db.attendances.find({"date": {"$gte": start, "$lte": end}}).to_list(None)

    enriched = await _enrich(records)
    by_employee = {}
    for r in enriched:
        eid = r["employeeId"]["_id"]
        if eid not in by_employee:
            by_employee[eid] = {
                "employeeId": r["employeeId"],
                "daysPresent": 0,
                "totalHours": 0.0,
            }
        if r["status"] == "present":
            by_employee[eid]["daysPresent"] += 1
        by_employee[eid]["totalHours"] += r["totalHours"]

    summary = list(by_employee.values())
    for s in summary:
        s["totalHours"] = round(s["totalHours"], 2)

    return {"month": month, "year": year, "records": enriched, "summary": summary}


@router.get("/employee/{eid}")
async def employee_report(eid: str, month: int = None, year: int = None):
    n = datetime.now(TZ)
    month = month or n.month
    year = year or n.year
    last_day = calendar.monthrange(year, month)[1]
    start = f"{year:04d}-{month:02d}-01"
    end = f"{year:04d}-{month:02d}-{last_day:02d}"

    records = await db.attendances.find({
        "employeeId": eid,
        "date": {"$gte": start, "$lte": end}
    }).to_list(None)

    enriched = await _enrich(records)
    days = sum(1 for r in enriched if r["status"] == "present")
    hours = round(sum(r["totalHours"] for r in enriched), 2)

    return {
        "month": month,
        "year": year,
        "records": enriched,
        "summary": {"daysPresent": days, "totalHours": hours},
    }

@router.get("/export")
async def export_report(
    type: str = Query("daily"),
    format: str = Query("xlsx"),
    date: str = None, month: int = None, year: int = None
):
    n = datetime.now(TZ)
    date = date or _today()
    month = month or n.month
    year = year or n.year
    last_day = calendar.monthrange(year, month)[1]
    filename = f"report_{date}_{month}_{year}"

    if type == "daily":
        records = await db.attendances.find({"date": date}).to_list(None)
        enriched = await _enrich(records)
        title = f"Daily Report - {date}"

        if format == "pdf":
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, title, ln=True, align="C")
            pdf.ln(5)
            pdf.set_font("Helvetica", "B", 7)
            headers = ["#", "Name", "In AM", "Out AM", "In PM", "Out PM", "Total"]
            col_w = [10, 40, 30, 30, 30, 30, 30]
            for i, h in enumerate(headers):
                pdf.cell(col_w[i], 8, h, border=1, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            for r in enriched:
                row = [
                    r["employeeId"]["employeeNumber"],
                    r["employeeId"]["fullName"][:16],
                    r["checkInAM"] or "-",
                    r["checkOutAM"] or "-",
                    r["checkInPM"] or "-",
                    r["checkOutPM"] or "-",
                    _hours_to_hms(r["totalHours"]),
                ]
                for i, val in enumerate(row):
                    pdf.cell(col_w[i], 6, val, border=1, align="C")
                pdf.ln()
            buf = BytesIO()
            pdf.output(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]
        ws.append(["#", "Name", "In AM", "Out AM", "In PM", "Out PM", "Total (hh:mm:ss)"])
        for r in enriched:
            ws.append([
                r["employeeId"]["employeeNumber"],
                r["employeeId"]["fullName"],
                r["checkInAM"] or "-",
                r["checkOutAM"] or "-",
                r["checkInPM"] or "-",
                r["checkOutPM"] or "-",
                _hours_to_hms(r["totalHours"]),
            ])
    else:
        start = f"{year:04d}-{month:02d}-01"
        end = f"{year:04d}-{month:02d}-{last_day:02d}"
        records = await db.attendances.find({"date": {"$gte": start, "$lte": end}}).to_list(None)
        enriched = await _enrich(records)

        by_employee = {}
        for r in enriched:
            eid = r["employeeId"]["_id"]
            if eid not in by_employee:
                by_employee[eid] = {
                    "employeeId": r["employeeId"],
                    "daysPresent": 0,
                    "totalHours": 0.0,
                }
            if r["status"] == "present":
                by_employee[eid]["daysPresent"] += 1
            by_employee[eid]["totalHours"] += r["totalHours"]

        summary = list(by_employee.values())
        for s in summary:
            s["totalHours"] = round(s["totalHours"], 2)

        title = f"Monthly Report - {year}-{month:02d}"

        if format == "pdf":
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, title, ln=True, align="C")
            pdf.ln(5)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(30, 8, "Number", border=1, align="C")
            pdf.cell(50, 8, "Name", border=1, align="C")
            pdf.cell(30, 8, "Days Present", border=1, align="C")
            pdf.cell(30, 8, "Total Hours", border=1, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 9)
            for s in summary:
                pdf.cell(30, 6, s["employeeId"]["employeeNumber"], border=1, align="C")
                pdf.cell(50, 6, s["employeeId"]["fullName"][:20], border=1)
                pdf.cell(30, 6, str(s["daysPresent"]), border=1, align="C")
                pdf.cell(30, 6, _hours_to_hms(s["totalHours"]), border=1, align="C")
                pdf.ln()
            buf = BytesIO()
            pdf.output(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]
        ws.append(["#", "Name", "Days Present", "Total Hours"])
        for s in summary:
            ws.append([
                s["employeeId"]["employeeNumber"],
                s["employeeId"]["fullName"],
                s["daysPresent"],
                _hours_to_hms(s["totalHours"]),
            ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )
